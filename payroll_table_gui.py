# نظام إدارة الرواتب - واجهة جدول محسّنة
# Payroll Management System - Enhanced Table GUI
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
from datetime import datetime
import shutil

# =============== صف الموظف Employee Class ===============
class Employee:
    def __init__(self, emp_id, name, base_salary, hours_per_day=8):
        self.emp_id = emp_id
        self.name = name
        self.base_salary = base_salary
        self.hours_per_day = hours_per_day
        
        # البيانات الشهرية - Monthly data
        self.absence_days = 0
        self.extra_days = 0
        self.extra_hours = 0
        self.tardiness_minutes = 0
        self.insurance = 0
        self.advance = 0
        self.advance_deduction = 0
        self.withdrawals = 0
    
    def calculate_net_salary(self):
        # حساب الراتب الصافي
        absence_deduction = (self.base_salary / 30) * self.absence_days
        extra_days_pay = (self.base_salary / 30) * self.extra_days
        extra_hours_pay = (self.base_salary / (30 * self.hours_per_day)) * self.extra_hours
        tardiness_deduction = (self.base_salary / (30 * self.hours_per_day * 60)) * self.tardiness_minutes
        
        net = (self.base_salary 
               + extra_days_pay 
               + extra_hours_pay 
               - absence_deduction 
               - tardiness_deduction 
               - self.insurance 
               - self.advance_deduction 
               - self.withdrawals)
        
        return round(net, 2)
    
    def to_dict(self):
        return {
            'emp_id': self.emp_id,
            'name': self.name,
            'base_salary': self.base_salary,
            'hours_per_day': self.hours_per_day,
            'absence_days': self.absence_days,
            'extra_days': self.extra_days,
            'extra_hours': self.extra_hours,
            'tardiness_minutes': self.tardiness_minutes,
            'insurance': self.insurance,
            'advance': self.advance,
            'advance_deduction': self.advance_deduction,
            'withdrawals': self.withdrawals
        }

# =============== تطبيق الواجهة الرئيسي Main App ===============
class PayrollTableApp:
    def __init__(self, root):
        self.root = root
        self.root.title("نظام إدارة الرواتب - Payroll Management System")
        self.root.geometry("1400x700")
        
        self.employees = []
        self.data_file = 'employees_data.json'
        
        # تحميل البيانات
        self.load_data()
        
        # إنشاء الواجهة
        self.create_widgets()
        
        # تحديث الجدول
        self.refresh_table()
    
    def create_widgets(self):
        # شريط العنوان
        title_frame = tk.Frame(self.root, bg='#2c3e50', height=60)
        title_frame.pack(fill=tk.X)
        
        title_label = tk.Label(title_frame, 
                              text="💼 نظام إدارة رواتب المصنع",
                              font=('Arial', 18, 'bold'),
                              bg='#2c3e50',
                              fg='white')
        title_label.pack(pady=15)
        
        # شريط الأزرار
        button_frame = tk.Frame(self.root, bg='#ecf0f1', height=50)
        button_frame.pack(fill=tk.X, padx=10, pady=5)
        
        buttons = [
            ("➕ إضافة موظف", self.add_employee),
            ("🗑️ حذف موظف", self.delete_employee),
            ("💾 حفظ", self.save_data),
            ("🔄 تحديث", self.refresh_table),
            ("📄 تصدير Excel", self.export_to_excel),
            ("📅 استيراد Excel", self.import_from_excel)
        ]
        
        for text, command in buttons:
            btn = tk.Button(button_frame, 
                           text=text,
                           command=command,
                           font=('Arial', 10, 'bold'),
                           bg='#3498db',
                           fg='white',
                           padx=15,
                           pady=5,
                           relief=tk.RAISED,
                           cursor='hand2')
            btn.pack(side=tk.LEFT, padx=5)
        
        # إطار الجدول
        table_frame = tk.Frame(self.root)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # شريط التمرير الأفقي
        h_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        # شريط التمرير العمودي
        v_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # الجدول (Treeview)
        columns = ('ID', 'Name', 'Base Salary', 'Hours', 'Absence Days', 
                   'Extra Days', 'Extra Hours', 'Tardiness Min', 
                   'Insurance', 'Advance', 'Advance Ded', 'Withdrawals', 'Net Salary')
        
        self.tree = ttk.Treeview(table_frame, 
                                columns=columns, 
                                show='headings',
                                xscrollcommand=h_scroll.set,
                                yscrollcommand=v_scroll.set,
                                height=20)
        
        h_scroll.config(command=self.tree.xview)
        v_scroll.config(command=self.tree.yview)
        
        # تحديد عناوين الأعمدة
        headers = ['الرقم', 'الاسم', 'الراتب الأساسي', 'الساعات', 'أيام الغياب', 
                   'أيام إضافية', 'ساعات إضافية', 'دقائق التأخير', 
                   'التأمين', 'السلفة', 'خصم السلفة', 'المسحوبات', 'الصافي']
        
        for col, header in zip(columns, headers):
            self.tree.heading(col, text=header)
            self.tree.column(col, width=100, anchor='center')
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # ربط النقر المزدوج للتحرير
        self.tree.bind('<Double-1>', self.on_double_click)
        
        # شريط الحالة
        status_frame = tk.Frame(self.root, bg='#34495e', height=30)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        self.status_label = tk.Label(status_frame,
                                     text="جاهز | Ready",
                                     font=('Arial', 9),
                                     bg='#34495e',
                                     fg='white')
        self.status_label.pack(side=tk.LEFT, padx=10)
    
    def refresh_table(self):
        # مسح الجدول
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # إضافة الموظفين
        for emp in self.employees:
            values = (emp.emp_id, emp.name, emp.base_salary, emp.hours_per_day,
                     emp.absence_days, emp.extra_days, emp.extra_hours,
                     emp.tardiness_minutes, emp.insurance, emp.advance,
                     emp.advance_deduction, emp.withdrawals, 
                     emp.calculate_net_salary())
            self.tree.insert('', tk.END, values=values, tags=('row',))
        
        self.status_label.config(text=f"تم تحديث الجدول - {len(self.employees)} موظف")
    
    def on_double_click(self, event):
        # الحصول على العنصر والعمود
        item = self.tree.identify('item', event.x, event.y)
        column = self.tree.identify_column(event.x)
        
        if not item or not column:
            return
        
        # الحصول على رقم العمود
        col_num = int(column.replace('#', '')) - 1
        
        # لا يمكن تحرير ID والاسم والصافي
        if col_num in [0, 1, 12]:
            messagebox.showinfo("تنبيه", "هذا الحقل غير قابل للتحرير")
            return
        
        # الحصول على القيمة الحالية
        current_value = self.tree.item(item, 'values')[col_num]
        
        # إنشاء نافذة التحرير
        self.edit_cell(item, col_num, current_value)
    
    def edit_cell(self, item, col_num, current_value):
        # نافذة التحرير
        edit_window = tk.Toplevel(self.root)
        edit_window.title("تحرير الحقل")
        edit_window.geometry("300x150")
        
        tk.Label(edit_window, text="أدخل القيمة الجديدة:", font=('Arial', 11)).pack(pady=10)
        
        entry = tk.Entry(edit_window, font=('Arial', 12), width=20)
        entry.insert(0, str(current_value))
        entry.pack(pady=10)
        entry.focus()
        
        def save_edit():
            new_value = entry.get()
            try:
                # تحويل إلى رقم
                new_value = float(new_value)
                
                # الحصول على الموظف
                emp_id = self.tree.item(item, 'values')[0]
                employee = next((e for e in self.employees if e.emp_id == emp_id), None)
                
                if employee:
                    # تحديث البيانات
                    field_map = {
                        2: 'base_salary',
                        3: 'hours_per_day',
                        4: 'absence_days',
                        5: 'extra_days',
                        6: 'extra_hours',
                        7: 'tardiness_minutes',
                        8: 'insurance',
                        9: 'advance',
                        10: 'advance_deduction',
                        11: 'withdrawals'
                    }
                    
                    if col_num in field_map:
                        setattr(employee, field_map[col_num], new_value)
                        self.refresh_table()
                        self.save_data()
                        edit_window.destroy()
                        messagebox.showinfo("نجاح", "تم تحديث البيانات بنجاح")
            except ValueError:
                messagebox.showerror("خطأ", "يرجى إدخال رقم صحيح")
        
        tk.Button(edit_window, 
                 text="حفظ",
                 command=save_edit,
                 bg='#27ae60',
                 fg='white',
                 font=('Arial', 11, 'bold'),
                 padx=20).pack(pady=10)
        
        entry.bind('<Return>', lambda e: save_edit())
    
    def add_employee(self):
        # نافذة إضافة موظف
        add_window = tk.Toplevel(self.root)
        add_window.title("إضافة موظف جديد")
        add_window.geometry("400x250")
        
        tk.Label(add_window, text="رقم الموظف:", font=('Arial', 10)).grid(row=0, column=0, padx=10, pady=10, sticky='e')
        emp_id_entry = tk.Entry(add_window, font=('Arial', 10))
        emp_id_entry.grid(row=0, column=1, padx=10, pady=10)
        
        tk.Label(add_window, text="الاسم:", font=('Arial', 10)).grid(row=1, column=0, padx=10, pady=10, sticky='e')
        name_entry = tk.Entry(add_window, font=('Arial', 10))
        name_entry.grid(row=1, column=1, padx=10, pady=10)
        
        tk.Label(add_window, text="الراتب الأساسي:", font=('Arial', 10)).grid(row=2, column=0, padx=10, pady=10, sticky='e')
        salary_entry = tk.Entry(add_window, font=('Arial', 10))
        salary_entry.grid(row=2, column=1, padx=10, pady=10)
        
        tk.Label(add_window, text="ساعات العمل:", font=('Arial', 10)).grid(row=3, column=0, padx=10, pady=10, sticky='e')
        hours_entry = tk.Entry(add_window, font=('Arial', 10))
        hours_entry.insert(0, "8")
        hours_entry.grid(row=3, column=1, padx=10, pady=10)
        
        def save_new_employee():
            try:
                emp_id = emp_id_entry.get()
                name = name_entry.get()
                salary = float(salary_entry.get())
                hours = float(hours_entry.get())
                
                if not emp_id or not name:
                    messagebox.showerror("خطأ", "يرجى ملء جميع الحقول")
                    return
                
                new_emp = Employee(emp_id, name, salary, hours)
                self.employees.append(new_emp)
                self.refresh_table()
                self.save_data()
                add_window.destroy()
                messagebox.showinfo("نجاح", "تمت إضافة الموظف بنجاح")
            except ValueError:
                messagebox.showerror("خطأ", "يرجى إدخال بيانات صحيحة")
        
        tk.Button(add_window, 
                 text="إضافة",
                 command=save_new_employee,
                 bg='#27ae60',
                 fg='white',
                 font=('Arial', 11, 'bold')).grid(row=4, column=0, columnspan=2, pady=20)
    
    def delete_employee(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("تحذير", "يرجى اختيار موظف للحذف")
            return
        
        emp_id = self.tree.item(selection[0], 'values')[0]
        
        if messagebox.askyesno("تأكيد", f"هل أنت متأكد من حذف الموظف {emp_id}؟"):
            self.employees = [e for e in self.employees if e.emp_id != emp_id]
            self.refresh_table()
            self.save_data()
            messagebox.showinfo("نجاح", "تم حذف الموظف بنجاح")
    
    def load_data(self):
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.employees = []
                    for emp_data in data:
                        emp = Employee(emp_data['emp_id'], 
                                     emp_data['name'],
                                     emp_data['base_salary'],
                                     emp_data.get('hours_per_day', 8))
                        emp.absence_days = emp_data.get('absence_days', 0)
                        emp.extra_days = emp_data.get('extra_days', 0)
                        emp.extra_hours = emp_data.get('extra_hours', 0)
                        emp.tardiness_minutes = emp_data.get('tardiness_minutes', 0)
                        emp.insurance = emp_data.get('insurance', 0)
                        emp.advance = emp_data.get('advance', 0)
                        emp.advance_deduction = emp_data.get('advance_deduction', 0)
                        emp.withdrawals = emp_data.get('withdrawals', 0)
                        self.employees.append(emp)
            except Exception as e:
                messagebox.showerror("خطأ", f"خطأ في تحميل البيانات: {e}")
    
    def save_data(self):
        try:
            # إنشاء نسخة احتياطية
            if os.path.exists(self.data_file):
                backup_file = f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{self.data_file}'
                shutil.copy(self.data_file, backup_file)
            
            # حفظ البيانات
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump([emp.to_dict() for emp in self.employees], f, ensure_ascii=False, indent=2)
            
            self.status_label.config(text="تم حفظ البيانات بنجاح")
        except Exception as e:
            messagebox.showerror("خطأ", f"خطأ في حفظ البيانات: {e}")
    
    def export_to_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "الرواتب"
            
            # العناوين
            headers = ['الرقم', 'الاسم', 'الراتب الأساسي', 'الساعات', 'أيام الغياب', 
                       'أيام إضافية', 'ساعات إضافية', 'دقائق التأخير',
                       'التأمين', 'السلفة', 'خصم السلفة', 'المسحوبات', 'الصافي']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # البيانات
            for row, emp in enumerate(self.employees, 2):
                ws.cell(row, 1, emp.emp_id)
                ws.cell(row, 2, emp.name)
                ws.cell(row, 3, emp.base_salary)
                ws.cell(row, 4, emp.hours_per_day)
                ws.cell(row, 5, emp.absence_days)
                ws.cell(row, 6, emp.extra_days)
                ws.cell(row, 7, emp.extra_hours)
                ws.cell(row, 8, emp.tardiness_minutes)
                ws.cell(row, 9, emp.insurance)
                ws.cell(row, 10, emp.advance)
                ws.cell(row, 11, emp.advance_deduction)
                ws.cell(row, 12, emp.withdrawals)
                ws.cell(row, 13, emp.calculate_net_salary())
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("ملف Excel", "*.xlsx")],
                initialfile=f"payroll_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
            if filename:
                wb.save(filename)
                messagebox.showinfo("نجاح", f"تم تصدير البيانات إلى: {filename}")
        except ImportError:
            messagebox.showerror("خطأ", "يرجى تثبيت openpyxl: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("خطأ", f"خطأ في التصدير: {e}")
    
    def import_from_excel(self):
        messagebox.showinfo("قريباً", "سيتم إضافة هذه الميزة قريباً")

# =============== تشغيل التطبيق ===============
if __name__ == '__main__':
    root = tk.Tk()
    app = PayrollTableApp(root)
    root.mainloop()
